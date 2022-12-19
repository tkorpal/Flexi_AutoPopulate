from imports import *
import openpyxl
from openpyxl import load_workbook 

wb = openpyxl.load_workbook("/home/ubuntu/FlexiAutoPopulate/oldFlexi.xlsx") 
worksheet = wb.active

for i in range(1, worksheet.max_row): 
    delete_filename(worksheet.cell(row=i, column=3).value)