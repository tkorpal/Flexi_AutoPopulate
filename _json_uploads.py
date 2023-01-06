from imports import *
import openpyxl
from openpyxl import load_workbook 

## Uploads Mulitple JSON files to AutoPopulate

work = openpyxl.load_workbook("/home/ubuntu/FlexiAutoPopulate/Test Banks.xlsx")
 
wb = work.active

def json_uploads():
    try:
        for x in range(1,wb.max_row):                              
            file = wb[f"A{x}"].value 
            year = file[:4]
            month = file[4:6]       
            account = file.split(' ')[1].split('-')[0]            
            json_file = f"{file.split('.')[0]}.json" 
            delete_file = file.split('.')[0].replace(' ', '_')
            # delete_file = f"{year}{month}_{file.split('.')[0]}"
            data = account_no_special(account, year)            
            if data:
                print(data) 
                delete_autopopulate_record(delete_file)
                source = f"{data[2]}{data[0]}/{year}/banking/{account}/.Autopopulate"             
                if os.path.exists(os.path.join(source, json_file)):
                    send_to_endpoint(source, json_file, data, month, year, 'holdings/', 'holdings')                
                    send_to_endpoint(source, json_file, data, month, year, 'transactions/', 'transactions')
                    update_to_null(file, 'processed')
                   
            
    except:
        pass            

    
if __name__ == '__main__':
    json_uploads()