from imports import *
 

import openpyxl
from openpyxl import load_workbook 

wb = openpyxl.load_workbook("/home/ubuntu/FlexiAutoPopulate/cleanup.xlsx")
worksheet = wb.active

for i in range(1, worksheet.max_row):    
    year = worksheet.cell(row=i, column=3).value[:4]   
    month =  worksheet.cell(row=i, column=3).value[4:7]
    json = (worksheet.cell(row=i, column=3).value).split('.')[0] + '.json'
    xlsx = (worksheet.cell(row=i, column=3).value).split('.')[0] + '.xlsx'
    account = (worksheet.cell(row=i, column=3).value).split('-')[0].split(' ')[1] 
    data = account_no_special(account, year)
    if data:
        destination = f"{data[2]}{data[0]}/{year}/Banking/{account}/.Autopopulate"
        if os.path.exists(os.path.join(destination, json)): 
            print(data)            
            update_database(data, worksheet.cell(row=i, column=3).value,  'holdings')
            update_database(data, worksheet.cell(row=i, column=3).value,  'transactions')
        else:
            print(data)             
            delete_filename(worksheet.cell(row=i, column=3).value)
    
     
    