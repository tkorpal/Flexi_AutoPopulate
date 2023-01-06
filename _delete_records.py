from imports import *
import openpyxl
from openpyxl import load_workbook 

## deletes record from a excel statement ##

wb = openpyxl.load_workbook("/home/ubuntu/FlexiAutoPopulate/oldFlexi.xlsx") 
worksheet = wb.active

for i in range(1, worksheet.max_row): 
    if worksheet.cell(row=i, column=3).value:
        print(worksheet.cell(row=i, column=3).value)
        delete_filename(worksheet.cell(row=i, column=3).value)